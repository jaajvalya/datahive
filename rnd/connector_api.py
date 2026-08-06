"""
Local R&D API — persist connector form payloads into MongoDB.

Uses MONGO_URI from repo-root `.env`. Default collections:
  connector_dtls       saved connections (passwords/keys encrypted)
  connection_log       connect / save outcomes
  query_log            Insights SQL executions
  glossary_upload_log  glossary uploads

Run (from this directory):
    python3 -m venv .venv
    .venv/bin/pip install -r requirements.txt
    .venv/bin/python connector_api.py

Listens on http://127.0.0.1:5055
"""
from __future__ import annotations

import json
import logging
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import Body, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, model_validator

_RND_DIR = Path(__file__).resolve().parent
_REPO_ROOT = _RND_DIR.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import asset_catalog  # noqa: E402
import connection_validator  # noqa: E402
import data_quality  # noqa: E402
import glossary_store  # noqa: E402
import mongo_store  # noqa: E402
import postgres_store  # noqa: E402
import snowflake_catalog  # noqa: E402

log = logging.getLogger("datahive.connector_api")

UPLOAD_DIR = _RND_DIR / "UPLOAD"
GLOSSARY_DIR = _RND_DIR / "GLOSSARY"
GLOSSARY_TEMPLATE_PATH = _RND_DIR / "templates" / "glossary_template.xlsx"
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_GLOSSARY_BYTES = 10 * 1024 * 1024
ALLOWED_UPLOAD_SUFFIXES = frozenset(
    {".csv", ".tsv", ".txt", ".xlsx", ".xls", ".json", ".parquet"}
)
ALLOWED_GLOSSARY_SUFFIXES = frozenset({".xlsx", ".xls", ".csv"})

mongo_store.load_repo_dotenv()
MONGO_URI = mongo_store.mongo_uri()
DB_NAME = mongo_store.database_name()
COLLECTION = mongo_store.connectors_collection_name()
CONNECTION_LOGS_COLLECTION = mongo_store.connection_logs_collection_name()
QUERY_LOGS_COLLECTION = mongo_store.query_logs_collection_name()
GLOSSARY_UPLOAD_LOG_COLLECTION = mongo_store.glossary_upload_logs_collection_name()
ASSET_GLOSSARY_COLLECTION = mongo_store.asset_glossary_collection_name()

_CONNECTION_LOG_PATH_PREFIXES = (
    "/api/connectors",
    "/api/connection-logs",
)


class SqlQueryIn(BaseModel):
    sql: str = Field(..., min_length=1)
    max_rows: int = Field(default=1000, ge=1, le=10_000)
    schema: str | None = None
    table: str | None = None
    connector_id: str | None = None


class DataQualityRunIn(BaseModel):
    connector_id: str = Field(..., min_length=1)
    schema: str = Field(..., min_length=1)
    tables: list[str] = Field(..., min_length=1, max_length=12)


class ConnectionLogIn(BaseModel):
    user: str | None = None
    message: str = Field(..., min_length=1)
    event: str = "connection.error"
    outcome: str = "failure"
    error_type: str | None = "client"
    context: dict[str, Any] | None = None

    @model_validator(mode="after")
    def _normalize_outcome(self) -> ConnectionLogIn:
        if self.outcome not in ("success", "failure"):
            self.outcome = "failure"
        if self.outcome == "success":
            self.error_type = None
        return self


def _resolve_user(request: Request | None, explicit: str | None = None) -> str:
    if explicit and explicit.strip():
        return explicit.strip()
    if request is not None:
        header_user = request.headers.get("X-DataHive-User")
        if header_user and header_user.strip():
            return header_user.strip()
        state_user = getattr(request.state, "user", None)
        if isinstance(state_user, str) and state_user.strip():
            return state_user.strip()
    return "unknown"


def _resolve_role(request: Request | None) -> str | None:
    if request is None:
        return None
    role = request.headers.get("X-DataHive-Role")
    return role.strip() if role and role.strip() else None


def log_connection_failure(
    user: str,
    message: str,
    *,
    event: str = "connection.error",
    error_type: str | None = "server",
    context: dict[str, Any] | None = None,
    http_status: int | None = None,
) -> None:
    try:
        mongo_store.log_connection_event(
            user,
            message,
            outcome="failure",
            event=event,
            error_type=error_type,
            context=context,
            http_status=http_status,
        )
    except RuntimeError as exc:
        log.error("connection_logs write failed: %s", exc)


def _ensure_upload_dir() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def _ensure_glossary_dir() -> None:
    GLOSSARY_DIR.mkdir(parents=True, exist_ok=True)


def _count_glossary_terms(path: Path) -> int | None:
    """Best-effort term/row count for Excel/CSV glossary uploads."""
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            with path.open("r", encoding="utf-8", errors="ignore") as fh:
                # header + data rows
                return max(0, sum(1 for _ in fh) - 1)
        if suffix in {".xlsx", ".xls"}:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            sheet = wb["Glossary"] if "Glossary" in wb.sheetnames else wb.active
            rows = 0
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    rows += 1
            wb.close()
            return rows
    except Exception as exc:  # noqa: BLE001
        log.warning("glossary term count failed for %s: %s", path.name, exc)
    return None


def _safe_stored_name(original: str) -> str:
    base = Path(original or "upload").name
    safe = re.sub(r"[^\w.\- ]", "_", base).strip() or "upload"
    return f"{uuid.uuid4().hex[:12]}_{safe}"


def _insert_connector_doc(doc: dict[str, Any]) -> dict[str, Any]:
    payload = dict(doc)
    payload.setdefault("connection_status", "connected")
    try:
        inserted_id = mongo_store.insert_connector_document(payload)
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    log_ok = True
    log_detail: str | None = None
    try:
        mongo_store.log_connection_event(
            str(payload.get("user") or "unknown"),
            f"Connection established — saved to {DB_NAME}.{COLLECTION}",
            outcome="success",
            event="connection.established",
            context={
                **mongo_store.connector_summary_context(payload),
                "connector_id": inserted_id,
                "connection_status": payload.get("connection_status"),
                "db": DB_NAME,
                "collection": COLLECTION,
            },
        )
    except RuntimeError as exc:
        log_ok = False
        log_detail = str(exc)
        log.error("connection_logs write after connector save failed: %s", exc)
    return {
        "ok": True,
        "id": inserted_id,
        "db": DB_NAME,
        "collection": COLLECTION,
        "connection_status": payload.get("connection_status"),
        "connection_log": log_ok,
        "connection_log_error": log_detail,
    }


async def _write_upload_file(upload: UploadFile, dest: Path) -> int:
    size = 0
    with dest.open("wb") as out:
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_UPLOAD_BYTES:
                out.close()
                dest.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=413,
                    detail=f"File exceeds {MAX_UPLOAD_BYTES // (1024 * 1024)} MB limit.",
                )
            out.write(chunk)
    return size


def _redacted_mongo_uri(uri: str) -> str:
    if "://" not in uri or "@" not in uri:
        return uri
    scheme, rest = uri.split("://", 1)
    _, host_part = rest.rsplit("@", 1)
    return f"{scheme}://***@{host_part}"


app = FastAPI(title="DataHive RND Connector Saver")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def attach_request_user(request: Request, call_next):
    request.state.user = request.headers.get("X-DataHive-User") or "unknown"
    return await call_next(request)


def _should_log_connection_path(path: str) -> bool:
    return any(path == p or path.startswith(p + "/") for p in _CONNECTION_LOG_PATH_PREFIXES)


@app.exception_handler(HTTPException)
async def connection_http_exception_handler(request: Request, exc: HTTPException):
    detail = exc.detail
    message = detail if isinstance(detail, str) else json.dumps(detail)
    if _should_log_connection_path(request.url.path):
        log_connection_failure(
            _resolve_user(request),
            message,
            event="connection.http_error",
            error_type="server",
            http_status=exc.status_code,
            context={
                "path": request.url.path,
                "method": request.method,
            },
        )
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


@app.exception_handler(Exception)
async def connection_unhandled_exception_handler(request: Request, exc: Exception):
    if _should_log_connection_path(request.url.path):
        log_connection_failure(
            _resolve_user(request),
            str(exc),
            event="connection.unhandled_error",
            error_type="server",
            http_status=500,
            context={
                "path": request.url.path,
                "method": request.method,
                "exception_type": type(exc).__name__,
            },
        )
    return JSONResponse(status_code=500, content={"detail": "Internal server error."})


def get_collection():
    return mongo_store.connectors_collection()


_RECENT_CONNECTOR_FIELDS = (
    "cloud",
    "connector_type",
    "display_name",
    "mode",
    "region",
    "account_id",
    "auth_type",
    "dataset_scope",
    "apis",
    "tenant_id",
    "resource_group",
    "access_key_id",
    "client_id",
    "role_arn",
    "file_name",
    "upload_format",
    "upload_notes",
    "connection_status",
    "credentials_encrypted",
    "credentials_keys",
    "user",
    "saved_at",
    "updated_at",
)


def _public_connector_item(doc: dict[str, Any]) -> dict[str, Any]:
    """Return a client-safe connector document (never includes secret values)."""
    item = dict(doc)
    if "_id" in item and "id" not in item:
        item["id"] = str(item.pop("_id"))
    elif "_id" in item:
        item.pop("_id", None)
    for key in (
        "api_key",
        "client_secret",
        "refresh_token",
        "secret_access_key",
        "service_account_json",
        "password",
        "private_key",
        "credentials_ciphertext",
    ):
        item.pop(key, None)
    return item


def _fetch_connectors(*, limit: int | None = None) -> list[dict[str, Any]]:
    """Return saved connectors (newest first). limit=None returns all."""
    projection = {field: 1 for field in _RECENT_CONNECTOR_FIELDS}
    projection["_id"] = 1
    cursor = get_collection().find({}, projection).sort(
        [("updated_at", -1), ("saved_at", -1), ("_id", -1)]
    )
    if limit is not None:
        cursor = cursor.limit(min(max(int(limit), 1), 500))
    return [_public_connector_item(doc) for doc in cursor]


def _fetch_recent_connectors(limit: int) -> list[dict[str, Any]]:
    return _fetch_connectors(limit=limit)


@app.get("/health")
def health(recent: int = 0) -> dict[str, Any]:
    try:
        get_collection()
        payload: dict[str, Any] = {
            "ok": True,
            "mongo": _redacted_mongo_uri(MONGO_URI),
            "db": DB_NAME,
            "collection": COLLECTION,
            "connection_logs_collection": CONNECTION_LOGS_COLLECTION,
            "query_logs_collection": QUERY_LOGS_COLLECTION,
            "glossary_upload_log_collection": GLOSSARY_UPLOAD_LOG_COLLECTION,
            "asset_glossary_collection": ASSET_GLOSSARY_COLLECTION,
            "sql_query_api": True,
            "query_log_api": True,
            "credentials_encrypted": True,
        }
        try:
            postgres_store.ping_postgres()
            payload["postgres"] = postgres_store.redacted_postgres_host()
            payload["postgres_ok"] = True
            kw = postgres_store.postgres_dsn_kwargs()
            payload["postgres_target"] = (
                f"{kw['user']}@{kw['host']}:{kw['port']}/{kw['dbname']}"
            )
            try:
                payload["asset_counts"] = postgres_store.catalog_counts()
                payload["asset_schemas"] = list(postgres_store.asset_schemas())
            except Exception as count_exc:  # noqa: BLE001
                payload["asset_counts_error"] = str(count_exc)
        except Exception as pg_exc:  # noqa: BLE001
            payload["postgres_ok"] = False
            payload["postgres_error"] = str(pg_exc)
        if recent > 0:
            try:
                payload["recent_connectors"] = _fetch_recent_connectors(recent)
            except Exception as recent_exc:  # noqa: BLE001
                payload["recent_connectors"] = []
                payload["recent_connectors_error"] = str(recent_exc)
        return payload
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"MongoDB unavailable: {exc}") from exc


@app.get("/api/assets/connectors")
def assets_connectors(request: Request) -> dict[str, Any]:
    """Connectors the current user may browse (privilege-aware)."""
    user = _resolve_user(request)
    role = _resolve_role(request)
    try:
        items = asset_catalog.list_accessible_connectors(user, role)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Connector list failed: {exc}") from exc
    return {
        "ok": True,
        "user": user,
        "admin": asset_catalog.user_is_admin(user, role),
        "count": len(items),
        "items": items,
    }


@app.get("/api/assets/catalog")
def assets_catalog(
    request: Request,
    connector_id: str | None = None,
) -> dict[str, Any]:
    """Unified assets across accessible connectors, optionally filtered to one."""
    user = _resolve_user(request)
    role = _resolve_role(request)
    try:
        return asset_catalog.build_catalog(
            user, role=role, connector_id=connector_id or "all"
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Asset catalog failed: {exc}") from exc


@app.get("/api/assets/relevant")
def assets_relevant(
    request: Request,
    tab: str = "recently_verified",
    type: str | None = None,
    connector_id: str | None = None,
) -> dict[str, Any]:
    if tab not in ("recently_verified", "my_drafts"):
        raise HTTPException(status_code=422, detail="invalid tab")
    user = _resolve_user(request)
    role = _resolve_role(request)
    try:
        if connector_id and connector_id not in ("all", asset_catalog.LOCAL_POSTGRES_ID):
            catalog = asset_catalog.build_catalog(
                user, role=role, connector_id=connector_id
            )
            items = catalog["items"]
            if type:
                items = [i for i in items if str(i.get("type") or "").lower() == type.lower()]
            return {
                "tab": tab,
                "items": items[:50],
                "counts": catalog.get("counts") or {},
                "connector_id": connector_id,
            }
        return postgres_store.relevant_assets(user, tab, type)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Assets relevant failed: {exc}") from exc


@app.get("/api/assets/search")
def assets_search(
    request: Request,
    q: str = "",
    limit: int = 10,
    offset: int = 0,
    connector_id: str | None = None,
) -> dict[str, Any]:
    user = _resolve_user(request)
    role = _resolve_role(request)
    query = (q or "").strip().lower()
    capped = min(max(limit, 1), 50)
    try:
        catalog = asset_catalog.build_catalog(
            user, role=role, connector_id=connector_id or "all"
        )
        items = catalog["items"]
        if query:
            items = [
                i
                for i in items
                if query in str(i.get("name") or "").lower()
                or query in str(i.get("schema") or "").lower()
                or query in str(i.get("crumb") or "").lower()
                or query in str(i.get("connector_name") or "").lower()
                or query in str(i.get("platform") or "").lower()
            ]
        page = items[offset : offset + capped]
        # Also include classic Postgres hits when browsing all / local postgres.
        if (not connector_id or connector_id in ("all", asset_catalog.LOCAL_POSTGRES_ID)) and query:
            try:
                pg = postgres_store.search_assets(user, q, limit=capped, offset=0)
                for hit in pg.get("items") or []:
                    hit = dict(hit)
                    hit.setdefault("connector_id", asset_catalog.LOCAL_POSTGRES_ID)
                    hit.setdefault("connector_name", "Local Postgres")
                    hit.setdefault("platform", "postgres")
                    page.append(hit)
            except Exception:  # noqa: BLE001
                pass
        # de-dupe by crumb+connector
        seen: set[str] = set()
        unique: list[dict[str, Any]] = []
        for hit in page:
            key = f"{hit.get('connector_id')}|{hit.get('crumb') or hit.get('name')}"
            if key in seen:
                continue
            seen.add(key)
            unique.append(hit)
        return {
            "q": q,
            "count": len(unique),
            "items": unique[:capped],
            "connector_id": connector_id or "all",
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Asset search failed: {exc}") from exc


@app.get("/api/assets/discover")
def assets_discover(
    request: Request,
    limit: int = 100,
    connector_id: str | None = None,
) -> dict[str, Any]:
    capped = min(max(limit, 1), 500)
    user = _resolve_user(request)
    role = _resolve_role(request)
    try:
        catalog = asset_catalog.build_catalog(
            user, role=role, connector_id=connector_id or "all"
        )
        return {
            "items": catalog["items"][:capped],
            "counts": catalog.get("counts") or {},
            "schemas": catalog.get("schemas") or [],
            "connectors": catalog.get("connectors") or [],
            "connector_id": catalog.get("selected_connector_id"),
            "asset_count": catalog.get("asset_count"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Asset discover failed: {exc}") from exc


@app.get("/api/assets/schemas")
def assets_schemas(
    request: Request,
    connector_id: str | None = None,
) -> dict[str, Any]:
    user = _resolve_user(request)
    role = _resolve_role(request)
    try:
        if not connector_id or connector_id == "all":
            catalog = asset_catalog.build_catalog(user, role=role, connector_id="all")
            return {
                "items": catalog.get("schemas") or [],
                "configured_schemas": list(postgres_store.asset_schemas()),
                "counts": catalog.get("counts") or {},
                "connectors": catalog.get("connectors") or [],
                "connector_count": catalog.get("connector_count"),
                "asset_count": catalog.get("asset_count"),
                "selected_connector_id": "all",
            }
        if connector_id == asset_catalog.LOCAL_POSTGRES_ID:
            return {
                "items": postgres_store.list_schemas(),
                "configured_schemas": list(postgres_store.asset_schemas()),
                "counts": postgres_store.catalog_counts(),
                "selected_connector_id": connector_id,
            }
        catalog = asset_catalog.build_catalog(
            user, role=role, connector_id=connector_id
        )
        return {
            "items": catalog.get("schemas") or [],
            "configured_schemas": [],
            "counts": catalog.get("counts") or {},
            "selected_connector_id": connector_id,
            "asset_count": catalog.get("asset_count"),
        }
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Asset schemas failed: {exc}") from exc


@app.get("/api/assets/counts")
def assets_counts(
    request: Request,
    connector_id: str | None = None,
) -> dict[str, Any]:
    user = _resolve_user(request)
    role = _resolve_role(request)
    try:
        catalog = asset_catalog.build_catalog(
            user, role=role, connector_id=connector_id or "all"
        )
        return {
            "counts": catalog.get("counts") or {},
            "connector_count": catalog.get("connector_count"),
            "asset_count": catalog.get("asset_count"),
            "selected_connector_id": catalog.get("selected_connector_id"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Asset counts failed: {exc}") from exc


def _asset_matches_schema(asset: dict[str, Any], schema: str) -> bool:
    """Match UI schema selection against catalog rows (incl. Snowflake DB.SCHEMA)."""
    wanted = (schema or "").strip().lower()
    if not wanted:
        return False
    candidates = {
        str(asset.get("schema") or "").lower(),
        str(asset.get("snowflake_schema") or "").lower(),
        str(asset.get("database") or "").lower(),
    }
    db = str(asset.get("database") or "").strip()
    sf_schema = str(asset.get("snowflake_schema") or "").strip()
    if db and sf_schema:
        candidates.add(f"{db}.{sf_schema}".lower())
    crumb = str(asset.get("crumb") or "")
    if crumb.count(".") >= 1:
        # DATABASE.SCHEMA.TABLE → DATABASE.SCHEMA
        parts = crumb.split(".")
        if len(parts) >= 2:
            candidates.add(f"{parts[0]}.{parts[1]}".lower())
    return wanted in candidates


def _catalog_table_items(
    catalog: dict[str, Any],
    schema: str,
    *,
    include_columns: bool = False,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for a in catalog.get("items") or []:
        if not _asset_matches_schema(a, schema):
            continue
        if a.get("name") == "snowflake_catalog_error":
            continue
        if str(a.get("type") or "") in {"Schema", "API", "Scope", "File"}:
            continue
        row = {
            "name": a["name"],
            "type": a.get("type") or "Table",
            "connector_id": a.get("connector_id"),
            "connector_name": a.get("connector_name"),
            "platform": a.get("platform"),
            "structure_supported": a.get("structure_supported"),
            "crumb": a.get("crumb"),
            "database": a.get("database"),
            "schema": a.get("schema"),
        }
        if include_columns and a.get("columns") is not None:
            row["columns"] = a.get("columns")
        items.append(row)
    return items


def _resolve_structure_connector(
    user: str,
    role: str | None,
    schema: str,
    table: str,
    connector_id: str | None,
) -> str:
    """Pick the connector that owns schema[.table] when Insights omits connector_id."""
    if connector_id and connector_id not in ("all",):
        return connector_id
    catalog = asset_catalog.build_catalog(user, role=role, connector_id="all")
    schema_assets = [
        a
        for a in catalog.get("items") or []
        if _asset_matches_schema(a, schema) and a.get("connector_id")
    ]
    matches = schema_assets
    if (table or "").strip():
        table_l = (table or "").strip().lower()
        exact = [
            a for a in schema_assets if str(a.get("name") or "").lower() == table_l
        ]
        if exact:
            matches = exact
    if matches:
        # Prefer Snowflake/live structure when multiple connectors share a name.
        matches.sort(
            key=lambda a: (
                0 if str(a.get("platform") or "").lower() == "snowflake" else 1,
                0 if a.get("structure_supported") else 1,
            )
        )
        return str(matches[0]["connector_id"])
    # Postgres medallion / local schemas
    if "." not in (schema or ""):
        return asset_catalog.LOCAL_POSTGRES_ID
    raise ValueError(
        f"Could not resolve connector for {schema}"
        + (f".{table}" if table else "")
        + ". Select a connector in Insights or Assets first."
    )


@app.get("/api/assets/tables")
def assets_tables(
    request: Request,
    schema: str,
    connector_id: str | None = None,
) -> dict[str, Any]:
    user = _resolve_user(request)
    role = _resolve_role(request)
    try:
        # Explicit Local Postgres only.
        if connector_id == asset_catalog.LOCAL_POSTGRES_ID:
            items = postgres_store.list_tables(schema)
            return {
                "schema": schema,
                "count": len(items),
                "items": items,
                "connector_id": asset_catalog.LOCAL_POSTGRES_ID,
                "platform": "postgres",
                "structure_supported": True,
            }

        # Default / all: unified catalog (Postgres + Snowflake + …).
        # Do NOT force Snowflake DB.SCHEMA labels through postgres_store.
        if not connector_id or connector_id == "all":
            catalog = asset_catalog.build_catalog(user, role=role, connector_id="all")
            items = _catalog_table_items(catalog, schema)
            if items:
                platforms = sorted(
                    {str(i.get("platform") or "") for i in items if i.get("platform")}
                )
                return {
                    "schema": schema,
                    "count": len(items),
                    "items": items,
                    "connector_id": "all",
                    "platforms": platforms,
                    "structure_supported": any(i.get("structure_supported") for i in items),
                }
            # Fallback: pure Postgres schema not yet in annotated catalog.
            if "." not in (schema or ""):
                try:
                    items = postgres_store.list_tables(schema)
                    return {
                        "schema": schema,
                        "count": len(items),
                        "items": items,
                        "connector_id": asset_catalog.LOCAL_POSTGRES_ID,
                        "platform": "postgres",
                        "structure_supported": True,
                    }
                except ValueError:
                    pass
            return {
                "schema": schema,
                "count": 0,
                "items": [],
                "connector_id": "all",
                "structure_supported": False,
                "note": f"No tables found for schema '{schema}' across connectors.",
            }

        catalog = asset_catalog.build_catalog(
            user, role=role, connector_id=connector_id
        )
        items = _catalog_table_items(catalog, schema, include_columns=True)
        note = None
        if not items:
            platform = None
            for c in catalog.get("connectors") or []:
                if c.get("id") == connector_id:
                    platform = c.get("platform") or c.get("cloud")
                    break
            if str(platform or "").lower() == "snowflake":
                note = (
                    f"No tables or views found in Snowflake schema '{schema}'. "
                    "The schema exists, but it is empty for this role."
                )
        return {
            "schema": schema,
            "count": len(items),
            "items": items,
            "connector_id": connector_id,
            "structure_supported": any(i.get("structure_supported") for i in items),
            "note": note,
        }
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Asset tables failed: {exc}") from exc


@app.get("/api/assets/structure")
def assets_structure(
    request: Request,
    schema: str,
    table: str,
    connector_id: str | None = None,
) -> dict[str, Any]:
    user = _resolve_user(request)
    role = _resolve_role(request)
    try:
        resolved_id = _resolve_structure_connector(
            user, role, schema, table, connector_id
        )
        if resolved_id == asset_catalog.LOCAL_POSTGRES_ID:
            structure = postgres_store.table_structure(schema, table)
            structure["connector_id"] = asset_catalog.LOCAL_POSTGRES_ID
            structure["connector_name"] = "Local Postgres"
            structure["platform"] = "postgres"
            return structure
        return asset_catalog.connector_structure(
            user,
            role=role,
            connector_id=resolved_id,
            schema=schema,
            table=table,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Asset structure failed: {exc}") from exc


def _looks_like_snowflake_sql(sql: str, schema: str | None) -> bool:
    """True for Snowflake 3-part refs or catalog schemas like SALES_DB.RAW."""
    if schema and "." in schema:
        return True
    text = sql or ""
    # "DB"."SCHEMA"."TABLE" or DB.SCHEMA.TABLE
    if re.search(
        r'(?:"[^"]+"|[A-Za-z_][\w$]*)\s*\.\s*(?:"[^"]+"|[A-Za-z_][\w$]*)\s*\.\s*(?:"[^"]+"|[A-Za-z_][\w$]*)',
        text,
    ):
        return True
    return False


def _resolve_snowflake_connector_id(
    user: str,
    role: str | None,
    schema: str | None,
    table: str | None,
    connector_id: str | None,
) -> str | None:
    """Resolve a real Snowflake connector_dtls id (skip glossary: stubs)."""
    candidates: list[str] = []
    if connector_id and connector_id not in ("all", asset_catalog.LOCAL_POSTGRES_ID):
        candidates.append(connector_id)
    if schema:
        try:
            resolved = _resolve_structure_connector(
                user, role, schema, table or "", None
            )
            if resolved and resolved not in candidates:
                candidates.append(resolved)
        except ValueError:
            pass
    try:
        by_platform = mongo_store.find_connector_by_platform("snowflake")
    except RuntimeError:
        by_platform = None
    if by_platform and by_platform.get("id"):
        candidates.append(str(by_platform["id"]))

    for cid in candidates:
        if not cid or str(cid).startswith("glossary:"):
            continue
        try:
            doc = mongo_store.get_connector_document(str(cid), with_secrets=False)
        except Exception:  # noqa: BLE001
            doc = None
        if not doc:
            continue
        cloud = str(doc.get("cloud") or doc.get("connector_type") or "").lower()
        if cloud == "snowflake":
            return str(cid)
    return None


@app.post("/api/sql/query")
def sql_query(body: SqlQueryIn, request: Request) -> dict[str, Any]:
    user = _resolve_user(request, None)
    role = _resolve_role(request)
    schema = (body.schema or "").strip() or None
    table = (body.table or "").strip() or None
    connector_id = (body.connector_id or "").strip() or None

    # Always prefer SQL-inferred 3-part names (DATABASE.SCHEMA / TABLE) when present.
    inferred_schema, inferred_table = postgres_store.infer_query_schema_table(body.sql)
    if inferred_schema and "." in inferred_schema:
        schema = inferred_schema
        table = inferred_table or table
    elif not schema or not table:
        schema = schema or inferred_schema
        table = table or inferred_table

    platform = "postgres"
    database = postgres_store.postgres_database_name()

    # Explicit non-snowflake connector → Postgres (or that connector's platform).
    explicit_platform = ""
    if connector_id and connector_id not in ("all", asset_catalog.LOCAL_POSTGRES_ID):
        if not str(connector_id).startswith("glossary:"):
            try:
                conn_doc = mongo_store.get_connector_document(
                    connector_id, with_secrets=False
                )
            except Exception:  # noqa: BLE001
                conn_doc = None
            if conn_doc:
                explicit_platform = str(
                    conn_doc.get("cloud") or conn_doc.get("connector_type") or ""
                ).lower()
                if explicit_platform == "snowflake":
                    platform = "snowflake"
                    database = (
                        (schema.split(".", 1)[0] if schema and "." in schema else None)
                        or str(conn_doc.get("dataset_scope") or "SALES_DB")
                    )
                elif explicit_platform and explicit_platform not in {
                    "postgres",
                    "postgresql",
                    "pg",
                    "local",
                }:
                    # Unknown cloud connectors currently have no SQL runner — keep Postgres
                    # only when the SQL is not clearly Snowflake-shaped.
                    pass

    if platform != "snowflake" and _looks_like_snowflake_sql(body.sql, schema):
        sf_id = _resolve_snowflake_connector_id(
            user, role, schema, table, connector_id
        )
        if sf_id:
            connector_id = sf_id
            platform = "snowflake"
            database = (
                schema.split(".", 1)[0]
                if schema and "." in schema
                else "SALES_DB"
            )

    query_start_time = datetime.now(timezone.utc)
    status = "success"
    error_message: str | None = None
    row_count: int | None = None
    result: dict[str, Any] | None = None
    try:
        if _looks_like_snowflake_sql(body.sql, schema) and platform != "snowflake":
            raise ValueError(
                "This SQL looks like a Snowflake query "
                f"({schema + '.' + table if schema and table else '3-part table name'}), "
                "but no Snowflake connector is available. "
                "Save/select the SFSALESDB connector in Insights."
            )
        if platform == "snowflake":
            if not connector_id:
                raise ValueError("Snowflake connector_id is required for this query.")
            doc = snowflake_catalog.load_connector_doc(connector_id)
            result = snowflake_catalog.execute_sql_query_for_doc(
                doc, body.sql, max_rows=body.max_rows
            )
        else:
            result = postgres_store.execute_sql_query(body.sql, max_rows=body.max_rows)
            result["platform"] = "postgres"
        row_count = result.get("row_count")
        result["connector_id"] = connector_id or asset_catalog.LOCAL_POSTGRES_ID
        return result
    except ValueError as exc:
        status = "failure"
        error_message = str(exc)
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        status = "failure"
        error_message = str(exc)
        label = "Snowflake" if platform == "snowflake" else "PostgreSQL"
        raise HTTPException(status_code=503, detail=f"{label} query failed: {exc}") from exc
    finally:
        query_end_time = datetime.now(timezone.utc)
        mongo_store.append_query_log(
            {
                "user": user,
                "source": "insights",
                "platform": platform,
                "connector_id": connector_id,
                "database": database,
                "schema": schema,
                "table": table,
                "query": body.sql,
                "max_rows": body.max_rows,
                "row_count": row_count,
                "truncated": (result or {}).get("truncated") if result else None,
                "query_start_time": query_start_time.isoformat(),
                "query_end_time": query_end_time.isoformat(),
                "duration_ms": int(
                    (query_end_time - query_start_time).total_seconds() * 1000
                ),
                "status": status,
                "error": error_message,
                "collection": QUERY_LOGS_COLLECTION,
            }
        )


@app.post("/api/data-quality/run")
def run_data_quality(body: DataQualityRunIn, request: Request) -> dict[str, Any]:
    """Profile selected tables and return DQ score, checks, and issue dashboard data."""
    user = _resolve_user(request)
    role = _resolve_role(request)
    connector_id = (body.connector_id or "").strip()
    schema = (body.schema or "").strip()
    tables = [str(t).strip() for t in (body.tables or []) if str(t).strip()]

    platform = "postgres"
    snowflake_doc: dict[str, Any] | None = None
    if connector_id and connector_id not in ("all", asset_catalog.LOCAL_POSTGRES_ID):
        if not str(connector_id).startswith("glossary:"):
            try:
                conn_doc = mongo_store.get_connector_document(
                    connector_id, with_secrets=False
                )
            except Exception:  # noqa: BLE001
                conn_doc = None
            if conn_doc:
                cloud = str(
                    conn_doc.get("cloud") or conn_doc.get("connector_type") or ""
                ).lower()
                if cloud == "snowflake":
                    platform = "snowflake"
                    snowflake_doc = snowflake_catalog.load_connector_doc(connector_id)

    def get_structure(sch: str, table: str) -> dict[str, Any]:
        if platform == "snowflake" and snowflake_doc is not None:
            return snowflake_catalog.table_structure_for_doc(snowflake_doc, sch, table)
        if connector_id in ("", "all", asset_catalog.LOCAL_POSTGRES_ID):
            return postgres_store.table_structure(sch.split(".")[-1], table)
        try:
            return asset_catalog.connector_structure(
                user,
                role=role,
                connector_id=connector_id,
                schema=sch,
                table=table,
            )
        except Exception:  # noqa: BLE001
            return postgres_store.table_structure(sch.split(".")[-1], table)

    def execute(sql: str) -> dict[str, Any]:
        if platform == "snowflake" and snowflake_doc is not None:
            return snowflake_catalog.execute_sql_query_for_doc(
                snowflake_doc, sql, max_rows=5000
            )
        return postgres_store.execute_sql_query(sql, max_rows=5000)

    try:
        return data_quality.run_data_quality(
            connector_id=connector_id or asset_catalog.LOCAL_POSTGRES_ID,
            schema=schema,
            tables=tables,
            platform=platform,
            get_structure=get_structure,
            execute=execute,
        )
    except data_quality.DataQualityError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=503, detail=f"Data quality run failed: {exc}"
        ) from exc


@app.get("/api/data-quality/logic")
def data_quality_logic() -> dict[str, Any]:
    return data_quality.score_logic_docs()


@app.post("/api/connection-logs")
def create_connection_log(body: ConnectionLogIn, request: Request) -> dict[str, bool]:
    try:
        mongo_store.log_connection_event(
            _resolve_user(request, body.user),
            body.message,
            outcome=body.outcome,  # validated success|failure
            event=body.event,
            error_type=body.error_type,
            context=body.context,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    return {"ok": True}


def _require_snowflake_connector(connector_id: str, user: str, role: str | None) -> dict[str, Any]:
    connectors = asset_catalog.list_accessible_connectors(user, role)
    conn = next((c for c in connectors if c["id"] == connector_id), None)
    if not conn:
        raise HTTPException(status_code=404, detail=f"Connector not found: {connector_id}")
    platform = str(conn.get("platform") or conn.get("cloud") or "").lower()
    if platform != "snowflake":
        raise HTTPException(status_code=422, detail="Connector is not a Snowflake connection.")
    return conn


@app.get("/api/snowflake/{connector_id}/stages")
def snowflake_list_stages(connector_id: str, request: Request) -> dict[str, Any]:
    user = _resolve_user(request)
    role = _resolve_role(request)
    _require_snowflake_connector(connector_id, user, role)
    try:
        stages = snowflake_catalog.list_stages_for_doc(
            snowflake_catalog.load_connector_doc(connector_id)
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Snowflake stage list failed: {exc}") from exc
    return {"ok": True, "connector_id": connector_id, "items": stages, "count": len(stages)}


@app.get("/api/snowflake/{connector_id}/stages/{stage_fqn:path}/files")
def snowflake_list_stage_files(
    connector_id: str,
    stage_fqn: str,
    request: Request,
    pattern: str = "",
) -> dict[str, Any]:
    user = _resolve_user(request)
    role = _resolve_role(request)
    _require_snowflake_connector(connector_id, user, role)
    try:
        files = snowflake_catalog.list_stage_files_for_doc(
            snowflake_catalog.load_connector_doc(connector_id),
            stage_fqn,
            pattern=pattern or "",
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except snowflake_catalog.StageAccessError as exc:
        return {
            "ok": True,
            "connector_id": connector_id,
            "stage_fqn": stage_fqn.lstrip("@"),
            "items": [],
            "count": 0,
            "exists": None,
            "visible": False,
            "reason": exc.reason,
            "note": str(exc),
            "grant_sql": snowflake_catalog.ensure_raw_stage_grant_sql(),
        }
    except Exception as exc:  # noqa: BLE001
        detail = str(exc)
        if "does not exist" in detail.lower() or "not authorized" in detail.lower():
            return {
                "ok": True,
                "connector_id": connector_id,
                "stage_fqn": stage_fqn.lstrip("@"),
                "items": [],
                "count": 0,
                "exists": None,
                "visible": False,
                "note": (
                    "Stage is not visible to this connector role (Snowflake returns "
                    "'does not exist or not authorized'). SALES_DB.RAW is owned by ACCOUNTADMIN; "
                    "SYSADMIN does not inherit those privileges. Grant READ/WRITE on the stage "
                    "— do not recreate it from DataHive."
                ),
                "grant_sql": snowflake_catalog.ensure_raw_stage_grant_sql(),
            }
        raise HTTPException(status_code=503, detail=f"Snowflake stage file list failed: {exc}") from exc
    return {
        "ok": True,
        "connector_id": connector_id,
        "stage_fqn": stage_fqn.lstrip("@"),
        "items": files,
        "count": len(files),
        "exists": True,
        "visible": True,
    }


@app.post("/api/snowflake/{connector_id}/stages/ensure-raw")
def snowflake_ensure_raw_stage(connector_id: str, request: Request) -> dict[str, Any]:
    """Create SALES_DB.RAW.RAW_STAGE (or scoped DB.RAW.RAW_STAGE) when permitted."""
    user = _resolve_user(request)
    role = _resolve_role(request)
    _require_snowflake_connector(connector_id, user, role)
    try:
        result = snowflake_catalog.ensure_raw_stage_for_doc(
            snowflake_catalog.load_connector_doc(connector_id)
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except PermissionError as exc:
        raise HTTPException(
            status_code=403,
            detail={
                "message": str(exc),
                "grant_sql": snowflake_catalog.ensure_raw_stage_grant_sql(),
                "workaround": "Use stage @~ (user stage) in ETL until RAW_STAGE is granted/created.",
            },
        ) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"Could not ensure RAW_STAGE: {exc}") from exc
    return {"ok": True, "connector_id": connector_id, **result}


@app.post("/api/connectors/test")
def test_connector(request: Request, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """
    Validate connector credentials with a live handshake before save.
    Accepts the same form payload as POST /api/connectors, or
    `{ "connector_id": "<id>" }` to re-test a saved connector.
    """
    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload")

    test_payload = dict(payload)
    connector_id = str(test_payload.pop("connector_id", "") or "").strip()
    if connector_id:
        try:
            doc = mongo_store.get_connector_document(connector_id, with_secrets=True)
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=503, detail=str(exc)) from exc
        if not doc:
            raise HTTPException(status_code=404, detail=f"Connector not found: {connector_id}")
        # Overlay any explicit fields from the request (non-secret overrides).
        for key, value in payload.items():
            if key == "connector_id":
                continue
            if value is not None and value != "":
                doc[key] = value
        test_payload = doc

    user = _resolve_user(request)
    cloud = str(test_payload.get("cloud") or test_payload.get("connector_type") or "")
    try:
        result = connection_validator.validate_connector(test_payload)
    except connection_validator.ConnectionValidationError as exc:
        log_connection_failure(
            user,
            str(exc),
            event="connection.validate_failed",
            error_type=exc.error_type or "auth",
            context={
                "cloud": cloud,
                "platform": exc.platform,
                "display_name": test_payload.get("display_name"),
                "account_id": test_payload.get("account_id"),
                "auth_type": test_payload.get("auth_type"),
                "connector_id": connector_id or None,
                "connection_status": "failed",
            },
        )
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        log_connection_failure(
            user,
            str(exc),
            event="connection.validate_error",
            error_type="server",
            context={
                "cloud": cloud,
                "display_name": test_payload.get("display_name"),
                "connector_id": connector_id or None,
            },
        )
        raise HTTPException(
            status_code=503,
            detail=f"Connection validation failed: {exc}",
        ) from exc

    mongo_store.log_connection_event(
        user,
        result.get("message") or "Connection validated",
        outcome="success",
        event="connection.validated",
        context={
            "cloud": cloud,
            "platform": result.get("platform"),
            "display_name": test_payload.get("display_name"),
            "account_id": test_payload.get("account_id"),
            "auth_type": test_payload.get("auth_type"),
            "connector_id": connector_id or None,
            "connection_status": "validated",
            "details": result.get("details") or {},
        },
    )
    return {
        "ok": True,
        "validated": True,
        "platform": result.get("platform"),
        "message": result.get("message"),
        "details": result.get("details") or {},
    }


@app.get("/api/connectors")
def list_connectors(limit: int | None = None) -> dict[str, Any]:
    """Return all saved connectors from MongoDB (newest first)."""
    try:
        items = _fetch_connectors(limit=limit)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"MongoDB read failed: {exc}") from exc
    return {
        "ok": True,
        "items": items,
        "count": len(items),
        "db": DB_NAME,
        "collection": COLLECTION,
    }


@app.post("/api/connectors")
def save_connector(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload")
    return _insert_connector_doc(dict(payload))


@app.get("/api/connectors/recent")
def list_recent_connectors(limit: int = 500) -> dict[str, Any]:
    """Compatibility alias for GET /api/connectors (returns all by default)."""
    try:
        items = _fetch_connectors(limit=limit)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"MongoDB read failed: {exc}") from exc
    return {
        "ok": True,
        "items": items,
        "count": len(items),
        "db": DB_NAME,
        "collection": COLLECTION,
    }


@app.get("/api/connectors/{connector_id}")
def get_connector(connector_id: str) -> dict[str, Any]:
    """Return a saved connector for editing. Secrets are never included."""
    try:
        doc = mongo_store.get_connector_document(connector_id, with_secrets=False)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    if not doc:
        raise HTTPException(status_code=404, detail=f"Connector not found: {connector_id}")
    return {"ok": True, "item": _public_connector_item(doc)}


@app.put("/api/connectors/{connector_id}")
def update_connector(
    connector_id: str,
    request: Request,
    payload: dict[str, Any] = Body(...),
) -> dict[str, Any]:
    """Update a saved connector. Blank secret fields keep existing values."""
    if not payload:
        raise HTTPException(status_code=400, detail="Empty payload")
    user = _resolve_user(request)
    try:
        updated = mongo_store.update_connector_document(connector_id, dict(payload))
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc

    try:
        mongo_store.log_connection_event(
            user,
            f"Connector updated — {DB_NAME}.{COLLECTION}",
            outcome="success",
            event="connection.updated",
            context={
                **mongo_store.connector_summary_context(updated),
                "connector_id": connector_id,
                "connection_status": updated.get("connection_status"),
                "db": DB_NAME,
                "collection": COLLECTION,
            },
        )
    except RuntimeError as exc:
        log.error("connection_logs write after connector update failed: %s", exc)

    return {
        "ok": True,
        "id": connector_id,
        "db": DB_NAME,
        "collection": COLLECTION,
        "connection_status": updated.get("connection_status"),
        "item": _public_connector_item(updated),
    }


@app.delete("/api/connectors/{connector_id}")
def delete_connector(connector_id: str, request: Request) -> dict[str, Any]:
    """Delete a saved connector from connector_dtls."""
    user = _resolve_user(request)
    summary: dict[str, Any] = {"connector_id": connector_id}
    try:
        existing = mongo_store.get_connector_document(connector_id, with_secrets=False)
        if existing:
            summary.update(mongo_store.connector_summary_context(existing))
        deleted = mongo_store.delete_connector_document(connector_id)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    if not deleted:
        raise HTTPException(status_code=404, detail=f"Connector not found: {connector_id}")

    try:
        mongo_store.log_connection_event(
            user,
            f"Connector deleted — {DB_NAME}.{COLLECTION}",
            outcome="success",
            event="connection.deleted",
            context={
                **summary,
                "db": DB_NAME,
                "collection": COLLECTION,
                "connection_status": "deleted",
            },
        )
    except RuntimeError as exc:
        log.error("connection_logs write after connector delete failed: %s", exc)

    return {"ok": True, "id": connector_id, "deleted": True}


@app.get("/api/connectors/{connector_id}/auth-ready")
def connector_auth_ready(connector_id: str) -> dict[str, Any]:
    """
    Server-side check that credentials can be decrypted for auth.
    Never returns secret values — only which fields are present.
    """
    try:
        creds = mongo_store.connector_credentials_for_auth(connector_id)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    return {
        "ok": True,
        "id": connector_id,
        "credentials_configured": bool(creds),
        "credential_fields": sorted(creds.keys()),
    }


@app.post("/api/connectors/upload")
async def save_connector_upload(
    file: UploadFile = File(...),
    metadata: str = Form(...),
) -> dict[str, Any]:
    try:
        meta = json.loads(metadata)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid metadata JSON.") from exc
    if not isinstance(meta, dict) or not meta:
        raise HTTPException(status_code=400, detail="Metadata must be a non-empty object.")

    original_name = file.filename or meta.get("file_name") or "upload"
    suffix = Path(original_name).suffix.lower()
    if suffix not in ALLOWED_UPLOAD_SUFFIXES:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type {suffix or '(none)'}.",
        )

    _ensure_upload_dir()
    stored_name = _safe_stored_name(original_name)
    dest = UPLOAD_DIR / stored_name
    bytes_written = await _write_upload_file(file, dest)

    doc = dict(meta)
    doc["mode"] = doc.get("mode") or "upload"
    doc["file_name"] = original_name
    doc["stored_file_name"] = stored_name
    doc["upload_relative_path"] = f"UPLOAD/{stored_name}"
    doc["file_size"] = bytes_written
    doc["file_type"] = file.content_type or doc.get("file_type") or ""

    result = _insert_connector_doc(doc)
    result["upload_relative_path"] = doc["upload_relative_path"]
    result["stored_file_name"] = stored_name
    return result


@app.get("/api/glossary/template")
def download_glossary_template() -> FileResponse:
    if not GLOSSARY_TEMPLATE_PATH.is_file():
        raise HTTPException(
            status_code=404,
            detail="glossary_template.xlsx was not found on the server.",
        )
    return FileResponse(
        path=str(GLOSSARY_TEMPLATE_PATH),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="glossary_template.xlsx",
    )


@app.get("/api/glossary/recent")
def recent_glossaries(limit: int = 20) -> dict[str, Any]:
    try:
        items = mongo_store.recent_glossary_documents(limit)
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    return {
        "ok": True,
        "items": items,
        "db": DB_NAME,
        "collection": GLOSSARY_UPLOAD_LOG_COLLECTION,
    }


@app.get("/api/glossary/terms")
def recent_glossary_terms(limit: int = 50) -> dict[str, Any]:
    """Unified asset glossary terms across AWS / Azure / GCP / Snowflake / Postgres."""
    try:
        items = mongo_store.recent_asset_glossary_terms(limit)
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    return {
        "ok": True,
        "items": items,
        "db": DB_NAME,
        "collection": ASSET_GLOSSARY_COLLECTION,
    }


@app.post("/api/glossary/upload")
async def upload_glossary(
    request: Request,
    file: UploadFile = File(...),
    notes: str = Form(""),
) -> dict[str, Any]:
    original_name = file.filename or "glossary.xlsx"
    suffix = Path(original_name).suffix.lower()
    if suffix not in ALLOWED_GLOSSARY_SUFFIXES:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Use .xlsx, .xls, or .csv.",
        )

    _ensure_glossary_dir()
    stored_name = _safe_stored_name(original_name)
    dest = GLOSSARY_DIR / stored_name
    bytes_written = await _write_upload_file(file, dest)
    if bytes_written > MAX_GLOSSARY_BYTES:
        dest.unlink(missing_ok=True)
        raise HTTPException(status_code=413, detail="Glossary file exceeds 10 MB limit.")

    term_count = _count_glossary_terms(dest)
    try:
        apply_result = glossary_store.apply_glossary_file(dest)
    except Exception as exc:  # noqa: BLE001
        dest.unlink(missing_ok=True)
        raise HTTPException(
            status_code=400,
            detail=f"Could not apply glossary to Assets metadata: {exc}",
        ) from exc

    user = _resolve_user(request)
    apply_summary = {
        "updated": apply_result.get("updated", 0),
        "registry_updated": apply_result.get("registry_updated", 0),
        "source_synced": apply_result.get("source_synced", 0),
        "skipped": apply_result.get("skipped", 0),
        "failed": apply_result.get("failed", 0),
        "rows_total": apply_result.get("rows_total", 0),
        "platforms": apply_result.get("platforms") or [],
        "errors": apply_result.get("errors") or [],
        "collection": apply_result.get("collection") or ASSET_GLOSSARY_COLLECTION,
    }
    doc = {
        "event": "glossary.upload",
        "outcome": "success" if apply_summary["failed"] == 0 else "partial",
        "file_name": original_name,
        "stored_file_name": stored_name,
        "upload_relative_path": f"GLOSSARY/{stored_name}",
        "file_size": bytes_written,
        "file_type": file.content_type or "",
        "notes": (notes or "").strip(),
        "term_count": term_count if term_count is not None else apply_result.get("rows_total"),
        "user": user,
        "kind": "glossary",
        "apply": apply_summary,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    try:
        inserted_id = mongo_store.insert_glossary_upload_log(doc)
    except RuntimeError as exc:
        # File + comments may already be applied; keep the file and report Mongo issue.
        raise HTTPException(status_code=503, detail=str(exc)) from exc

    return {
        "ok": True,
        "id": inserted_id,
        "db": DB_NAME,
        "collection": GLOSSARY_UPLOAD_LOG_COLLECTION,
        "asset_glossary_collection": ASSET_GLOSSARY_COLLECTION,
        "file_name": original_name,
        "stored_file_name": stored_name,
        "upload_relative_path": doc["upload_relative_path"],
        "file_size": bytes_written,
        "term_count": doc["term_count"],
        "apply": doc["apply"],
        "updates": apply_result.get("updates") or [],
    }


@app.get("/api/glossary/files/{stored_name}")
def download_glossary_upload(stored_name: str) -> FileResponse:
    safe = Path(stored_name).name
    path = GLOSSARY_DIR / safe
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Glossary file not found.")
    return FileResponse(
        path=str(path),
        filename=safe.split("_", 1)[-1] if "_" in safe else safe,
        media_type="application/octet-stream",
    )


# Serve the R&D UI over HTTP so the browser can reach this API.
# Opening main.html as file:// often yields "Failed to fetch" (blocked localhost).
_IMAGES_DIR = _REPO_ROOT / "images"
_UI_ASSET_RE = re.compile(
    r"^[A-Za-z0-9._-]+\.(html|js|css|svg|png|jpg|jpeg|webp|ico|map|woff2?|ttf)$"
)


@app.get("/", include_in_schema=False)
def serve_ui_index() -> FileResponse:
    path = _RND_DIR / "main.html"
    if not path.is_file():
        raise HTTPException(status_code=404, detail="UI not found.")
    return FileResponse(path, media_type="text/html")


@app.get("/main.html", include_in_schema=False)
def serve_ui_main() -> FileResponse:
    return serve_ui_index()


@app.get("/{filename}", include_in_schema=False)
def serve_ui_asset(filename: str) -> FileResponse:
    reserved = {"docs", "redoc", "openapi.json", "health", "favicon.ico"}
    if filename in reserved or not _UI_ASSET_RE.match(filename):
        raise HTTPException(status_code=404, detail="Not found.")
    path = (_RND_DIR / filename).resolve()
    if not str(path).startswith(str(_RND_DIR.resolve())) or not path.is_file():
        raise HTTPException(status_code=404, detail="Not found.")
    return FileResponse(path)


if _IMAGES_DIR.is_dir():
    app.mount("/images", StaticFiles(directory=str(_IMAGES_DIR)), name="images")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=5055, log_level="info")
