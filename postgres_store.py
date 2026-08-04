"""PostgreSQL access for searchable / discoverable data assets (repo `.env` POSTGRES_*)."""
from __future__ import annotations

import json
import logging
import os
import re
from contextlib import contextmanager
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterator

from mongo_store import load_repo_dotenv

_log = logging.getLogger("datahive.postgres_store")

_REPO_ROOT = Path(__file__).resolve().parent

MEDALLION_ASSET_SCHEMAS = ("bronze", "silver", "gold")
DEFAULT_ASSET_SCHEMAS = MEDALLION_ASSET_SCHEMAS

_SCHEMA_STATEMENTS = (
    """
    CREATE TABLE IF NOT EXISTS assets (
        id SERIAL PRIMARY KEY,
        owner TEXT NOT NULL,
        name TEXT NOT NULL,
        type TEXT NOT NULL,
        crumb TEXT,
        edited TEXT,
        verified BOOLEAN NOT NULL DEFAULT FALSE,
        warning BOOLEAN NOT NULL DEFAULT FALSE,
        tab TEXT NOT NULL DEFAULT 'recently_verified'
    )
    """,
    "CREATE INDEX IF NOT EXISTS assets_owner_tab_idx ON assets (owner, tab)",
)


def _dotenv_file_values() -> dict[str, str]:
    """Read repo-root `.env`; file values beat os.environ for POSTGRES_* (setdefault misses updates)."""
    out: dict[str, str] = {}
    path = _REPO_ROOT / ".env"
    if not path.is_file():
        return out
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        if key.startswith("POSTGRES_"):
            out[key] = value.strip().strip('"').strip("'")
    return out


def _pg_setting(name: str, default: str) -> str:
    load_repo_dotenv()
    file_vals = _dotenv_file_values()
    if name in file_vals and file_vals[name] != "":
        return file_vals[name]
    return os.environ.get(name, default)


def asset_schemas() -> tuple[str, ...]:
    """Asset catalog always includes bronze, silver, gold; `.env` may add more schemas."""
    raw = _pg_setting("POSTGRES_ASSET_SCHEMAS", "").strip()
    extras = [p.strip() for p in raw.split(",") if p.strip()] if raw else []
    ordered: list[str] = []
    seen: set[str] = set()
    for name in (*MEDALLION_ASSET_SCHEMAS, *extras):
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(name)
    return tuple(ordered) if ordered else MEDALLION_ASSET_SCHEMAS


def postgres_dsn_kwargs() -> dict[str, Any]:
    load_repo_dotenv()
    return {
        "host": _pg_setting("POSTGRES_HOST", "localhost"),
        "port": int(_pg_setting("POSTGRES_PORT", "5432")),
        "dbname": _pg_setting("POSTGRES_DATABASE", "datahive"),
        "user": _pg_setting("POSTGRES_USER", "postgres"),
        "password": _pg_setting("POSTGRES_PASSWORD", ""),
    }


def postgres_database_name() -> str:
    return postgres_dsn_kwargs()["dbname"]


def postgres_conninfo() -> str | None:
    """Optional URI override. Uses POSTGRES_CONNINFO or POSTGRES_URI only (not DATABASE_URL)."""
    for key in ("POSTGRES_CONNINFO", "POSTGRES_URI"):
        val = _pg_setting(key, "").strip()
        if val:
            return val
    return None


def postgres_connect():
    """Open a psycopg connection using repo `.env` POSTGRES_* settings."""
    import psycopg
    from psycopg.conninfo import make_conninfo

    explicit = postgres_conninfo()
    if explicit:
        return psycopg.connect(explicit, connect_timeout=5)
    kw = postgres_dsn_kwargs()
    return psycopg.connect(
        make_conninfo(
            host=kw["host"],
            port=kw["port"],
            dbname=kw["dbname"],
            user=kw["user"],
            password=kw["password"],
        ),
        connect_timeout=5,
    )


def redacted_postgres_host() -> str:
    kw = postgres_dsn_kwargs()
    schemas = ", ".join(asset_schemas())
    return f"{kw['user']}@{kw['host']}:{kw['port']}/{kw['dbname']} (schemas: {schemas})"


@contextmanager
def postgres_connection() -> Iterator[Any]:
    conn = postgres_connect()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def ping_postgres() -> None:
    with postgres_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1")


def ensure_assets_schema() -> None:
    with postgres_connection() as conn:
        with conn.cursor() as cur:
            for stmt in _SCHEMA_STATEMENTS:
                cur.execute(stmt)
            cur.execute("SELECT COUNT(*) FROM assets")
            count = int(cur.fetchone()[0])
            if count == 0:
                seed = [
                    (
                        "Admin",
                        "Revenue by Region (draft)",
                        "Query",
                        "Query › Revenue by Region",
                        "edited yesterday",
                        False,
                        False,
                        "my_drafts",
                    ),
                ]
                cur.executemany(
                    """
                    INSERT INTO assets (owner, name, type, crumb, edited, verified, warning, tab)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    seed,
                )


def _is_system_schema(schema: str) -> bool:
    name = (schema or "").strip().lower()
    return (not name) or name.startswith("pg_") or name == "information_schema"


def _allowed_schema_name(schema: str) -> bool:
    """Allow configured medallion/extra schemas and any non-system schema in Postgres."""
    name = (schema or "").strip()
    if _is_system_schema(name):
        return False
    if name.lower() in {s.lower() for s in MEDALLION_ASSET_SCHEMAS}:
        return True
    allowed = asset_schemas()
    if name in allowed or name.lower() in {a.lower() for a in allowed}:
        return True
    # Live catalog: any user schema present in the connected database.
    try:
        with postgres_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT 1
                    FROM pg_catalog.pg_namespace
                    WHERE nspname = %s OR lower(nspname) = lower(%s)
                    LIMIT 1
                    """,
                    (name, name),
                )
                return cur.fetchone() is not None
    except Exception:  # noqa: BLE001
        return False


def _resolve_namespace(cur: Any, schema: str) -> str:
    cur.execute(
        """
        SELECT nspname
        FROM pg_catalog.pg_namespace
        WHERE nspname = %s OR lower(nspname) = lower(%s)
        ORDER BY CASE WHEN nspname = %s THEN 0 ELSE 1 END
        LIMIT 1
        """,
        (schema, schema, schema),
    )
    row = cur.fetchone()
    if row is None:
        raise ValueError(f"Schema '{schema}' was not found in PostgreSQL.")
    return str(row[0])


def _list_user_namespaces(cur: Any) -> list[str]:
    """All non-system schemas in the connected database (pg_catalog)."""
    cur.execute(
        """
        SELECT nspname
        FROM pg_catalog.pg_namespace
        WHERE nspname NOT LIKE 'pg\\_%' ESCAPE '\\'
          AND nspname <> 'information_schema'
        ORDER BY
          CASE lower(nspname)
            WHEN 'bronze' THEN 0
            WHEN 'silver' THEN 1
            WHEN 'gold' THEN 2
            WHEN 'public' THEN 3
            ELSE 10
          END,
          nspname
        """
    )
    return [str(r[0]) for r in cur.fetchall()]


def _resolved_asset_schemas(cur: Any) -> list[str]:
    """Configured schemas that exist, then any other user schemas from Postgres."""
    configured_existing: list[str] = []
    for name in asset_schemas():
        try:
            configured_existing.append(_resolve_namespace(cur, name))
        except ValueError:
            continue

    seen = {s.lower() for s in configured_existing}
    discovered = configured_existing[:]
    for nsp in _list_user_namespaces(cur):
        if nsp.lower() in seen:
            continue
        seen.add(nsp.lower())
        discovered.append(nsp)
    return discovered


def _relation_type_label(relkind: str) -> str:
    if relkind in ("v", "m"):
        return "View"
    if relkind == "f":
        return "Foreign Table"
    return "Table"


def _list_relations_in_namespace(cur: Any, nspname: str) -> list[tuple[str, str, str]]:
    """Return (name, asset_type, relkind) for user relations in one schema."""
    cur.execute(
        """
        SELECT c.relname, c.relkind
        FROM pg_catalog.pg_class c
        JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
        WHERE n.nspname = %s
          AND c.relkind IN ('r', 'p', 'v', 'm', 'f')
          AND NOT c.relispartition
        ORDER BY c.relname
        """,
        (nspname,),
    )
    out: list[tuple[str, str, str]] = []
    for name, relkind in cur.fetchall():
        out.append((str(name), _relation_type_label(str(relkind)), str(relkind)))
    return out


def _catalog_counts() -> dict[str, int]:
    """All = column count; View / Table = relation counts in asset schemas (pg_catalog)."""
    with postgres_connection() as conn:
        with conn.cursor() as cur:
            resolved = _resolved_asset_schemas(cur)
            if not resolved:
                return {"All": 0, "View": 0, "Table": 0}

            cur.execute(
                """
                SELECT COUNT(*)
                FROM information_schema.columns
                WHERE table_schema = ANY(%s)
                """,
                (resolved,),
            )
            fields_n = int(cur.fetchone()[0])

            tables_n = 0
            views_n = 0
            for nsp in resolved:
                for _name, asset_type, _kind in _list_relations_in_namespace(cur, nsp):
                    if asset_type == "View":
                        views_n += 1
                    elif asset_type in ("Table", "Foreign Table"):
                        tables_n += 1

    return {"All": fields_n, "View": views_n, "Table": tables_n}


def catalog_counts() -> dict[str, int]:
    """Public wrapper for chip counts (All / View / Table)."""
    return _catalog_counts()


def _catalog_assets() -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    with postgres_connection() as conn:
        with conn.cursor() as cur:
            for nsp in _resolved_asset_schemas(cur):
                for name, asset_type, _kind in _list_relations_in_namespace(cur, nsp):
                    items.append(
                        {
                            "name": name,
                            "type": asset_type,
                            "crumb": f"{nsp} › {name}",
                            "edited": "catalog snapshot",
                            "verified": True,
                            "warning": False,
                            "source": "catalog",
                            "schema": nsp,
                        }
                    )
    return items


def _table_assets(owner: str, tab: str) -> list[dict[str, Any]]:
    sql = """
        SELECT name, type, crumb, edited, verified, warning
        FROM assets
        WHERE owner = %s AND tab = %s
        ORDER BY id DESC
    """
    with postgres_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (owner, tab))
            rows = cur.fetchall()
    return [
        {
            "name": r[0],
            "type": r[1],
            "crumb": r[2] or r[1],
            "edited": r[3] or "",
            "verified": bool(r[4]),
            "warning": bool(r[5]),
            "source": "assets",
        }
        for r in rows
    ]


def _apply_type_filter(
    items: list[dict[str, Any]], asset_type: str | None
) -> list[dict[str, Any]]:
    if not asset_type or asset_type == "All":
        return items
    return [i for i in items if i.get("type") == asset_type]


def relevant_assets(
    owner: str,
    tab: str = "recently_verified",
    asset_type: str | None = None,
) -> dict[str, Any]:
    ensure_assets_schema()
    counts = _catalog_counts()
    if tab == "my_drafts":
        items = _table_assets(owner, "my_drafts")
    else:
        items = _catalog_assets()
    items = _apply_type_filter(items, asset_type)
    return {
        "tab": tab,
        "type": asset_type,
        "schemas": list(asset_schemas()),
        "counts": counts,
        "items": items,
    }


def search_assets(
    owner: str,
    query: str,
    *,
    limit: int = 10,
    offset: int = 0,
) -> dict[str, Any]:
    ensure_assets_schema()
    q = (query or "").strip()
    limit = max(1, min(limit, 50))
    offset = max(0, offset)
    items: list[dict[str, Any]] = []
    if not q:
        return {"query": q, "items": items}

    like = f"%{q}%"
    with postgres_connection() as conn:
        with conn.cursor() as cur:
            schemas = _resolved_asset_schemas(cur)
            if not schemas:
                return {"query": q, "items": items}

            cur.execute(
                """
                SELECT n.nspname, c.relname, c.relkind
                FROM pg_catalog.pg_class c
                JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
                WHERE n.nspname = ANY(%s)
                  AND c.relkind IN ('r', 'p', 'v', 'm', 'f')
                  AND NOT c.relispartition
                  AND (c.relname ILIKE %s OR n.nspname ILIKE %s)
                ORDER BY n.nspname, c.relname
                LIMIT %s OFFSET %s
                """,
                (schemas, like, like, limit, offset),
            )
            for schema, name, relkind in cur.fetchall():
                asset_type = _relation_type_label(str(relkind))
                if asset_type == "Foreign Table":
                    asset_type = "Table"
                items.append({"name": name, "type": asset_type, "schema": schema})

            if len(items) < limit:
                cur.execute(
                    """
                    SELECT table_schema, table_name, column_name
                    FROM information_schema.columns
                    WHERE table_schema = ANY(%s)
                      AND (column_name ILIKE %s OR table_name ILIKE %s)
                    ORDER BY table_schema, table_name, ordinal_position
                    LIMIT %s
                    """,
                    (schemas, like, like, limit - len(items)),
                )
                seen = {(i["name"], i["type"], i.get("schema")) for i in items}
                for schema, table, column in cur.fetchall():
                    label = f"{table}.{column}"
                    key = (label, "Column", schema)
                    if key in seen:
                        continue
                    items.append({"name": label, "type": "Column", "schema": schema})
                    seen.add(key)

            if len(items) < limit:
                cur.execute(
                    """
                    SELECT DISTINCT name, type
                    FROM assets
                    WHERE owner = %s AND (name ILIKE %s OR type ILIKE %s OR crumb ILIKE %s)
                    ORDER BY name
                    LIMIT %s
                    """,
                    (owner, like, like, like, limit - len(items)),
                )
                seen = {(i["name"], i["type"]) for i in items}
                for name, typ in cur.fetchall():
                    key = (name, typ)
                    if key in seen:
                        continue
                    items.append({"name": name, "type": typ})
                    seen.add(key)

    return {"query": q, "items": items[:limit]}


def discover_assets(owner: str, *, limit: int = 100) -> dict[str, Any]:
    ensure_assets_schema()
    _ = owner
    _ = limit
    items = _catalog_assets()
    return {"items": items, "schemas": list(asset_schemas()), "counts": _catalog_counts()}


def list_schemas() -> list[str]:
    """Schemas from the connected PostgreSQL database (configured first, then other user schemas)."""
    configured = list(asset_schemas())
    with postgres_connection() as conn:
        with conn.cursor() as cur:
            found = _resolved_asset_schemas(cur)
    # De-dupe while preserving order
    seen: set[str] = set()
    ordered: list[str] = []
    for s in found:
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(s)
    if ordered:
        return ordered
    # Namespace missing in DB — still expose configured names for the UI
    return configured


def list_tables(schema: str) -> list[dict[str, Any]]:
    """All tables/views in one schema (pg_catalog — matches what you see in Postgres)."""
    schema = (schema or "").strip()
    if not _allowed_schema_name(schema):
        raise ValueError(f"Schema '{schema}' is not in the configured asset schemas.")
    with postgres_connection() as conn:
        with conn.cursor() as cur:
            nsp = _resolve_namespace(cur, schema)
            rels = _list_relations_in_namespace(cur, nsp)
    return [{"name": name, "type": asset_type} for name, asset_type, _kind in rels]


def table_structure(schema: str, table: str) -> dict[str, Any]:
    """Column-level structure (name, type, nullable, default, primary key) for one table/view."""
    schema = (schema or "").strip()
    table = (table or "").strip()
    if not _allowed_schema_name(schema):
        raise ValueError(f"Schema '{schema}' is not in the configured asset schemas.")

    with postgres_connection() as conn:
        with conn.cursor() as cur:
            nsp = _resolve_namespace(cur, schema)
            cur.execute(
                """
                SELECT c.relkind
                FROM pg_catalog.pg_class c
                JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
                WHERE n.nspname = %s AND c.relname = %s
                """,
                (nsp, table),
            )
            row = cur.fetchone()
            if row is None:
                raise ValueError(f"Table '{nsp}.{table}' was not found.")
            table_type = _relation_type_label(str(row[0]))

            cur.execute(
                """
                SELECT column_name, data_type, is_nullable, column_default,
                       character_maximum_length, numeric_precision, numeric_scale,
                       ordinal_position
                FROM information_schema.columns
                WHERE table_schema = %s AND table_name = %s
                ORDER BY ordinal_position
                """,
                (nsp, table),
            )
            cols = cur.fetchall()

            cur.execute(
                """
                SELECT kcu.column_name
                FROM information_schema.table_constraints tc
                JOIN information_schema.key_column_usage kcu
                  ON tc.constraint_name = kcu.constraint_name
                 AND tc.table_schema = kcu.table_schema
                WHERE tc.constraint_type = 'PRIMARY KEY'
                  AND tc.table_schema = %s AND tc.table_name = %s
                """,
                (nsp, table),
            )
            pk_columns = {r[0] for r in cur.fetchall()}

            # Column comments (JSON business metadata when present).
            cur.execute(
                """
                SELECT a.attname, pg_catalog.col_description(a.attrelid, a.attnum)
                FROM pg_catalog.pg_attribute a
                JOIN pg_catalog.pg_class c ON c.oid = a.attrelid
                JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
                WHERE n.nspname = %s
                  AND c.relname = %s
                  AND a.attnum > 0
                  AND NOT a.attisdropped
                """,
                (nsp, table),
            )
            comments = {str(r[0]): r[1] for r in cur.fetchall()}

            cur.execute(
                """
                SELECT pg_catalog.obj_description(c.oid, 'pg_class')
                FROM pg_catalog.pg_class c
                JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
                WHERE n.nspname = %s AND c.relname = %s
                """,
                (nsp, table),
            )
            table_comment_row = cur.fetchone()
            table_comment = table_comment_row[0] if table_comment_row else None

    columns: list[dict[str, Any]] = []
    for name, data_type, nullable, default, char_len, num_prec, num_scale, position in cols:
        type_display = data_type
        if char_len:
            type_display = f"{data_type}({char_len})"
        elif num_prec is not None and data_type in ("numeric", "decimal"):
            type_display = (
                f"{data_type}({num_prec},{num_scale})" if num_scale else f"{data_type}({num_prec})"
            )
        raw_comment = comments.get(name)
        metadata = _parse_column_metadata(raw_comment)
        columns.append(
            {
                "position": position,
                "name": name,
                "type": type_display,
                "nullable": nullable == "YES",
                "default": default,
                "primary_key": name in pk_columns,
                "comment": raw_comment,
                "metadata": metadata,
            }
        )

    return {
        "schema": nsp,
        "table": table,
        "table_type": table_type,
        "comment": table_comment,
        "metadata": _parse_column_metadata(table_comment),
        "columns": columns,
    }


def _parse_column_metadata(comment: Any) -> dict[str, Any] | None:
    """Parse JSON business metadata from a Postgres comment; return None if absent/invalid."""
    if comment is None:
        return None
    text = str(comment).strip()
    if not text:
        return None
    if text[0] not in "{[":
        return {"description": text}
    try:
        parsed = json.loads(text)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {"description": text}
    return parsed if isinstance(parsed, dict) else {"value": parsed}


GLOSSARY_ID_FIELDS = ("connection", "database", "schema", "table", "column")
GLOSSARY_META_FIELDS = (
    "business_name",
    "description",
    "business_definition",
    "classification",
    "sensitivity",
    "source_system",
    "owner",
    "steward",
)
_GLOSSARY_HEADER_ALIASES = {
    "business_defintion": "business_definition",
    "business definition": "business_definition",
    "business_name": "business_name",
    "source system": "source_system",
    "conn": "connection",
    "db": "database",
    "column_name": "column",
    "col": "column",
}


def _normalize_glossary_header(raw: Any) -> str:
    text = str(raw or "").strip().lower().replace("-", "_")
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return _GLOSSARY_HEADER_ALIASES.get(text, text)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_glossary_rows(path: Path) -> list[dict[str, str]]:
    """Parse glossary xlsx/csv into normalized row dicts."""
    suffix = path.suffix.lower()
    rows_out: list[dict[str, str]] = []

    if suffix == ".csv":
        import csv

        with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as fh:
            reader = csv.reader(fh)
            try:
                header = next(reader)
            except StopIteration:
                return []
            keys = [_normalize_glossary_header(h) for h in header]
            for raw in reader:
                if not any(_cell_str(v) for v in raw):
                    continue
                item = {keys[i]: _cell_str(raw[i]) if i < len(raw) else "" for i in range(len(keys))}
                rows_out.append(item)
        return rows_out

    if suffix not in {".xlsx", ".xls"}:
        raise ValueError(f"Unsupported glossary file type: {suffix or '(none)'}")

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb["Glossary"] if "Glossary" in wb.sheetnames else wb.active
        it = sheet.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            return []
        keys = [_normalize_glossary_header(h) for h in header]
        for raw in it:
            if raw is None or not any(_cell_str(v) for v in raw):
                continue
            item = {
                keys[i]: _cell_str(raw[i]) if i < len(raw) else ""
                for i in range(len(keys))
            }
            rows_out.append(item)
    finally:
        wb.close()
    return rows_out


def _postgres_connect_to(dbname: str):
    """Open a psycopg connection to a specific database using repo POSTGRES_* settings."""
    import psycopg
    from psycopg.conninfo import make_conninfo

    target = (dbname or "").strip() or postgres_database_name()
    explicit = postgres_conninfo()
    if explicit:
        # Override dbname in URI when possible; otherwise fall back to kwargs.
        try:
            return psycopg.connect(explicit, dbname=target, connect_timeout=5)
        except TypeError:
            pass
    kw = postgres_dsn_kwargs()
    return psycopg.connect(
        make_conninfo(
            host=kw["host"],
            port=kw["port"],
            dbname=target,
            user=kw["user"],
            password=kw["password"],
        ),
        connect_timeout=5,
    )


def _column_exists(cur: Any, schema: str, table: str, column: str) -> tuple[str, str, str] | None:
    """Return resolved (schema, table, column) names if the column exists."""
    cur.execute(
        """
        SELECT n.nspname, c.relname, a.attname
        FROM pg_catalog.pg_attribute a
        JOIN pg_catalog.pg_class c ON c.oid = a.attrelid
        JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
        WHERE (n.nspname = %s OR lower(n.nspname) = lower(%s))
          AND (c.relname = %s OR lower(c.relname) = lower(%s))
          AND (a.attname = %s OR lower(a.attname) = lower(%s))
          AND a.attnum > 0
          AND NOT a.attisdropped
        ORDER BY
          CASE WHEN n.nspname = %s THEN 0 ELSE 1 END,
          CASE WHEN c.relname = %s THEN 0 ELSE 1 END,
          CASE WHEN a.attname = %s THEN 0 ELSE 1 END
        LIMIT 1
        """,
        (schema, schema, table, table, column, column, schema, table, column),
    )
    row = cur.fetchone()
    if not row:
        return None
    return str(row[0]), str(row[1]), str(row[2])


def _suggest_column_target(
    cur: Any, schema: str, table: str, column: str
) -> str | None:
    """Best-effort suggestion when an identifier does not resolve exactly."""
    import difflib

    cur.execute(
        """
        SELECT nspname FROM pg_catalog.pg_namespace
        WHERE nspname NOT LIKE 'pg\\_%' ESCAPE '\\'
          AND nspname <> 'information_schema'
        """
    )
    schemas = [str(r[0]) for r in cur.fetchall()]
    schema_match = difflib.get_close_matches(schema, schemas, n=1, cutoff=0.6)
    use_schema = schema_match[0] if schema_match else schema

    cur.execute(
        """
        SELECT c.relname
        FROM pg_catalog.pg_class c
        JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
        WHERE (n.nspname = %s OR lower(n.nspname) = lower(%s))
          AND c.relkind IN ('r', 'p', 'v', 'm', 'f')
        """,
        (use_schema, use_schema),
    )
    tables = [str(r[0]) for r in cur.fetchall()]
    table_match = difflib.get_close_matches(table, tables, n=1, cutoff=0.6)
    use_table = table_match[0] if table_match else table

    cur.execute(
        """
        SELECT a.attname
        FROM pg_catalog.pg_attribute a
        JOIN pg_catalog.pg_class c ON c.oid = a.attrelid
        JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
        WHERE (n.nspname = %s OR lower(n.nspname) = lower(%s))
          AND (c.relname = %s OR lower(c.relname) = lower(%s))
          AND a.attnum > 0 AND NOT a.attisdropped
        """,
        (use_schema, use_schema, use_table, use_table),
    )
    columns = [str(r[0]) for r in cur.fetchall()]
    column_match = difflib.get_close_matches(column, columns, n=1, cutoff=0.6)

    if schema_match or table_match or column_match:
        return (
            f"Did you mean "
            f"{use_schema}.{use_table}."
            f"{(column_match[0] if column_match else column)}?"
        )
    if schemas:
        return "Available schemas: " + ", ".join(schemas[:12])
    return None


def _set_column_comment(cur: Any, schema: str, table: str, column: str, comment: str) -> None:
    from psycopg import sql

    cur.execute(
        sql.SQL("COMMENT ON COLUMN {}.{}.{} IS {}").format(
            sql.Identifier(schema),
            sql.Identifier(table),
            sql.Identifier(column),
            sql.Literal(comment),
        )
    )


def apply_glossary_file(path: Path | str) -> dict[str, Any]:
    """
    Apply glossary spreadsheet rows as PostgreSQL column comments.

    Identifier columns: connection, database, schema, table, column
    Metadata columns: business_name, description, business_definition,
    classification, sensitivity, source_system, owner, steward
    """
    file_path = Path(path)
    rows = _read_glossary_rows(file_path)
    if not rows:
        return {
            "rows_total": 0,
            "updated": 0,
            "skipped": 0,
            "failed": 0,
            "errors": ["No data rows found in glossary file."],
            "updates": [],
        }

    # Validate required headers exist in at least one usable form.
    present = set()
    for row in rows:
        present.update(k for k, v in row.items() if k)
    missing_headers = [h for h in ("schema", "table", "column") if h not in present]
    if missing_headers:
        return {
            "rows_total": len(rows),
            "updated": 0,
            "skipped": 0,
            "failed": len(rows),
            "errors": [
                "Missing required columns: "
                + ", ".join(missing_headers)
                + ". Expected identifiers: connection, database, schema, table, column."
            ],
            "updates": [],
        }

    updated = 0
    skipped = 0
    failed = 0
    errors: list[str] = []
    updates: list[dict[str, Any]] = []
    # Cache connections per database name
    conn_cache: dict[str, Any] = {}

    try:
        for idx, row in enumerate(rows, start=2):  # 1-based sheet rows; header is row 1
            connection = row.get("connection", "")
            database = row.get("database", "") or postgres_database_name()
            schema = row.get("schema", "")
            table = row.get("table", "")
            column = row.get("column", "")

            if not schema or not table or not column:
                skipped += 1
                errors.append(f"Row {idx}: skipped — schema/table/column are required.")
                continue

            meta: dict[str, Any] = {}
            for key in GLOSSARY_META_FIELDS:
                val = row.get(key, "")
                if val != "":
                    meta[key] = val
            if connection:
                meta["connection"] = connection
            if database:
                meta["database"] = database
            if not meta:
                skipped += 1
                errors.append(f"Row {idx}: skipped — no metadata fields to apply.")
                continue

            try:
                if database not in conn_cache:
                    conn_cache[database] = _postgres_connect_to(database)
                    conn_cache[database].autocommit = True
                conn = conn_cache[database]
                with conn.cursor() as cur:
                    resolved = _column_exists(cur, schema, table, column)
                    if not resolved:
                        failed += 1
                        hint = _suggest_column_target(cur, schema, table, column)
                        msg = (
                            f"Row {idx}: column not found — "
                            f"{database}.{schema}.{table}.{column}"
                        )
                        if hint:
                            msg += f" ({hint})"
                        errors.append(msg)
                        continue
                    nsp, rel, att = resolved
                    # Preserve unknown keys from an existing JSON comment.
                    cur.execute(
                        """
                        SELECT pg_catalog.col_description(c.oid, a.attnum)
                        FROM pg_catalog.pg_attribute a
                        JOIN pg_catalog.pg_class c ON c.oid = a.attrelid
                        JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
                        WHERE n.nspname = %s AND c.relname = %s AND a.attname = %s
                        """,
                        (nsp, rel, att),
                    )
                    existing_raw = cur.fetchone()
                    existing = _parse_column_metadata(existing_raw[0] if existing_raw else None) or {}
                    if not isinstance(existing, dict):
                        existing = {}
                    merged = dict(existing)
                    merged.update(meta)
                    comment = json.dumps(merged, ensure_ascii=False)
                    _set_column_comment(cur, nsp, rel, att, comment)
                updated += 1
                updates.append(
                    {
                        "row": idx,
                        "connection": connection,
                        "database": database,
                        "schema": nsp,
                        "table": rel,
                        "column": att,
                    }
                )
            except Exception as exc:  # noqa: BLE001
                failed += 1
                errors.append(
                    f"Row {idx}: failed — {database}.{schema}.{table}.{column}: {exc}"
                )
    finally:
        for conn in conn_cache.values():
            try:
                conn.close()
            except Exception:  # noqa: BLE001
                pass

    return {
        "rows_total": len(rows),
        "updated": updated,
        "skipped": skipped,
        "failed": failed,
        "errors": errors[:50],
        "updates": updates[:100],
    }


_SQL_COMMENT_BLOCK = re.compile(r"/\*.*?\*/", re.DOTALL)
_SQL_COMMENT_LINE = re.compile(r"--[^\n]*")
_FORBIDDEN_SQL = re.compile(
    r"\b("
    r"INSERT|UPDATE|DELETE|DROP|TRUNCATE|ALTER|CREATE|GRANT|REVOKE|"
    r"COPY|CALL|DO|EXECUTE|MERGE|REPLACE|VACUUM|ANALYZE|CLUSTER|REINDEX|"
    r"COMMENT|LOCK|UNLOCK|DISCARD|LISTEN|NOTIFY|PREPARE|DEALLOCATE|"
    r"REFRESH|REASSIGN|SECURITY|SET\s+ROLE|RESET\s+ROLE"
    r")\b",
    re.IGNORECASE,
)
_ALLOWED_SQL_START = frozenset({"SELECT", "WITH", "EXPLAIN", "TABLE", "VALUES"})


def _strip_sql_comments(sql: str) -> str:
    without_block = _SQL_COMMENT_BLOCK.sub(" ", sql)
    return _SQL_COMMENT_LINE.sub(" ", without_block)


_QUALIFIED_TABLE_REF = re.compile(
    r"\b(?:FROM|JOIN)\s+"
    r'(?:"([^"]+)"|([a-zA-Z_][\w$]*))\s*\.\s*'
    r'(?:"([^"]+)"|([a-zA-Z_][\w$]*))',
    re.IGNORECASE,
)
_UNQUALIFIED_FROM_REF = re.compile(
    r'\bFROM\s+(?:"([^"]+)"|([a-zA-Z_][\w$]*))(?:\s|$)',
    re.IGNORECASE,
)


def infer_query_schema_table(sql: str) -> tuple[str | None, str | None]:
    """Best-effort schema/table from the first FROM/JOIN reference in SQL."""
    normalized = _strip_sql_comments(sql.strip().rstrip(";"))
    match = _QUALIFIED_TABLE_REF.search(normalized)
    if match:
        schema = match.group(1) or match.group(2)
        table = match.group(3) or match.group(4)
        return schema, table
    match = _UNQUALIFIED_FROM_REF.search(normalized)
    if match:
        return None, match.group(1) or match.group(2)
    return None, None


def _assert_readonly_sql(sql: str) -> str:
    raw = sql.strip()
    if not raw:
        raise ValueError("Query is empty.")
    body = raw.rstrip(";").strip()
    if ";" in body:
        raise ValueError("Only a single SQL statement is allowed.")
    normalized = _strip_sql_comments(body)
    if _FORBIDDEN_SQL.search(normalized):
        raise ValueError("Only read-only SELECT queries are allowed.")
    match = re.match(r"^\s*(\w+)", normalized)
    if not match:
        raise ValueError("Invalid SQL.")
    keyword = match.group(1).upper()
    if keyword not in _ALLOWED_SQL_START:
        raise ValueError(
            f"Statement type '{keyword}' is not allowed. Use SELECT (or WITH … SELECT)."
        )
    return body


def _json_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (bytes, bytearray, memoryview)):
        return bytes(value).hex()
    return str(value)


def execute_sql_query(sql: str, *, max_rows: int = 1000) -> dict[str, Any]:
    """Run one read-only SQL statement; returns column names and row values."""
    statement = _assert_readonly_sql(sql)
    capped = min(max(int(max_rows), 1), 10_000)

    with postgres_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("BEGIN READ ONLY")
            cur.execute("SET LOCAL statement_timeout = '60000'")
            cur.execute(statement)
            if cur.description is None:
                return {
                    "columns": [],
                    "rows": [],
                    "row_count": 0,
                    "truncated": False,
                    "max_rows": capped,
                }
            columns = [col.name for col in cur.description]
            fetched = cur.fetchmany(capped + 1)
            truncated = len(fetched) > capped
            if truncated:
                fetched = fetched[:capped]
            rows = [[_json_cell(cell) for cell in row] for row in fetched]
            return {
                "columns": columns,
                "rows": rows,
                "row_count": len(rows),
                "truncated": truncated,
                "max_rows": capped,
            }
